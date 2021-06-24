# excel_funcs.py

"""
all functions relevant to excel, spreadsheets, and data storage
"""


def extract_courses(csv_fname):
    """
    extracts all course/course number combinations as a list of tuples
    :param csv_fname: csv file containing all departments of interest and their respective courses
    :return: courses: list of all department-course number tuples
    """

    import csv  # import csv library to parse and manipulate .csv files

    courses = []    # initialize empty list
    with open(csv_fname, newline='', encoding='utf-8-sig') as f:    # open csv file
        cf = csv.DictReader(f)
        for row in cf:  # loop through rows
            for column_name, value in row.items():  # loop through columns
                if value.strip():
                    courses.append((column_name, value))
    courses.sort()  # now holds tuples of all course/course number combinations
    return courses


def init_excel_storage(data_fname, json_list):
    """
    initializes an excel spreadsheet for CARATS to store data to
    #:param quarter: current quarter
    #:param data_fext: type of file extension (.xls, .xlsm, .xlsx)
    :param data_fname: name of file in which data will be stored
    :param json_list: list of all json objects
    :return: saves a populated workbook of requested extension
    """

    import openpyxl                         # import openpyxl library to manage excel sheets (install via pip)
    from json_funcs import json_data_parse  # import json data parser to parse json objects in json list

    # data_fname = "CARATS_Qtr_" + str(quarter) + data_fext   # title excel spreadsheet

    wb = openpyxl.Workbook()    # create workbook

    # create and title sheets in workbook
    base_title = "All Courses "
    data_collected = ["Space Left", "% Space Left", "Total Space"]   # sheet titles
    ws_space = wb.active                                             # set active sheet as first sheet
    ws_space.title = base_title + data_collected[0]                  # space left
    ws_pspace = wb.create_sheet(title=base_title+data_collected[1])  # % space left
    ws_tspace = wb.create_sheet(title=base_title+data_collected[2])  # total space left
    sheets = [ws_space, ws_pspace, ws_tspace]                        # put sheets in a list iterate over them

    staticHeaders = ["Enrollment Code", "Course Name", "Professor(s)", "Days and Times"]    # static column titles

    # fill in first row static header values
    for ws in sheets:                                      # loop through all sheets
        for (hdr_idx, hdr) in enumerate(staticHeaders):    # loop through all headers
            ws.cell(column=hdr_idx + 1, row=1, value=hdr)  # write in header value in cell

    # correction vector to adjust new row population
    corr = [0] * len(staticHeaders)  # initialize correction for max row addition
    corr[0] = corr[0] + 1            # update correction vector

    # parse through json objects
    for jsonObj in json_list:               # json list input, output is excel sheet
        statics = json_data_parse(jsonObj)  # list of all relevant course data corresponding to json object

        # write values in corresponding cell
        for ws in sheets:   # loop through all sheets
            for (rowIdx, staticRow) in enumerate(statics):  # loop through all course data lists in course list
                for colIdx in range(0, 4):   # loop through all 4 static columns
                    ws.cell(column=colIdx + 1, row=ws.max_row + corr[colIdx], value=staticRow[colIdx])  # write

    wb.save(filename=data_fname)    # save workbook to specified name in specified format


def time_excel_storage(data_fname, curr_time_str, json_list):
    """
    populates data spreadsheet for CARATS
    :param data_fname: name of file in which data will be stored
    :param curr_time_str: current time (MM/DD/YYYY HH:MM
    :param json_list: list of all json objects
    :return: saves updated workbook
    """

    # imports
    import openpyxl                         # import openpyxl library to manage excel sheets (install via pip)
    from json_funcs import json_data_parse  # import json data parser to parse json objects in json list

    wb = openpyxl.load_workbook(data_fname)  # load worksheet to update

    # input current date/time in headers
    for ws in wb.worksheets:    # loop through all worksheets
        ws.cell(column=ws.max_column + 1, row=1).value = curr_time_str  # insert current time string value

    # parse through json objects
    for json_obj in json_list:                # loop through all json objects in json list
        temporal = json_data_parse(json_obj)  # list of all relevant course data corresponding to json object

        # update values
        for (course_idx, course) in enumerate(temporal):  # loop through all course data lists in course list
            enrlCd = course[0]                            # enrollment code

            # search for row to update by comparing enrollment codes
            for (ws_idx, ws) in enumerate(wb):
                update_row = 0  # set default to impossible value
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):  # loop through rows
                    for curr_cell in row:               # loop through each cell in the row
                        if curr_cell.value == enrlCd:   # compare enrollment code values
                            update_row = curr_cell.row  # change row to update
                            break                       # break out of for loop

                # check nonexistent enrollment codes
                if update_row == 0:  # what to do if enrollment code doesn't exist
                    update_row = ws.max_row + 1         # add a new row
                    for static_labels in range(0, 4):   # loop through static data
                        ws.cell(column=static_labels + 1, row=update_row, value=course[static_labels])  # write

                # update cell with data from corresponding sheet
                ws.cell(column=ws.max_column, row=update_row).value = course[ws_idx + 4]    # write data value

    wb.save(filename=data_fname)  # save workbook to specified name in specified format
