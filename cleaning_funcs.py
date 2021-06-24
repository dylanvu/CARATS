import openpyxl
from openpyxl.styles import Font, numbers, Alignment
import win32com.client as win32
from pathlib import Path

# Abhiram: Please note that pywin32 must be installed via the terminal!


def condense_labs(data_fname, condense_fname):
    # This function takes in an input path to the CARATS output worksheet and saves the excel. For example, a valid
    # input may be path = "CARATS_Qtr_20211.xlsx" if this file is in the same folder as the script.

    # Resource for style using openpyxl: https://openpyxl.readthedocs.io/en/stable/styles.html
    # Please note that if you reopen a formatted excel sheet and save it, all formatting will be lost!

    wb_obj = openpyxl.load_workbook(data_fname, keep_vba=True)

    # When importing this function into the main script, comment out the above lines.

    # TODO: Optimize code. Put dictionary creation outside of the loop, all max row/column calculations outside

    # Iterate through the sheets
    print("Condensing courses.")
    for sheet in range(0, 3):
        wb_obj.active = sheet

        sheet_obj = wb_obj.active

        # The value of max_cols will correspond to the cell with no other data in it
        max_cols = sheet_obj.max_column
        # Formatting the document's static first row (the header) by making it bold
        for col in range(1, max_cols + 1):
            sheet_obj.cell(row=1, column=col).font = Font(bold=True)
        # Freeze the static headers and all course information
        sheet_obj.freeze_panes = "E5"

        # TODO: We can put the dictionary creation outside of the loop to optimize the code
        # Creating a dictionary of a row number as the key and the lab course information as the value
        lab_courses = {}

        # For every course row
        for row in range(1, sheet_obj.max_row+1):
            # A hacky way to reformat percentages? TODO: Maybe put this percentage format elsewhere?
            # If the current sheet is the percentages sheet
            if sheet == 1:
                # For all columns of data
                for col in range(5, max_cols + 1):
                    sheet_obj.cell(row=row, column=col).number_format = numbers.FORMAT_PERCENTAGE

            name_cell_obj = sheet_obj.cell(row=row, column=2)
            course_name_split = sheet_obj.cell(row=row, column=2).value.split()
            # If the course num ends in L (like CHEM 1AL), put into dictionary with the row as the key and info as value.
            if "L" in course_name_split[len(course_name_split) - 1]:
                professor_cell_obj = sheet_obj.cell(row=row, column=3)
                time_cell_obj = sheet_obj.cell(row=row, column=4)
                lab_name = name_cell_obj.value
                professor = professor_cell_obj.value  # What if like the lab professors change and enrollment code stays the same?
                time = time_cell_obj.value
                lab_courses[row] = tuple([lab_name, professor, time])

        # Log the total number of rows in the Excel document for later. This row index represents the first entirely blank row
        max_row = row

        # Start counting how many unique courses there are
        unique_courses = []

        lab_courses_values = list(lab_courses.values())
        for course in range(0, len(lab_courses_values)):
            if lab_courses_values[course] not in unique_courses:
                unique_courses.append(lab_courses_values[course])

        lab_groupings = {}
        # For each unique course
        for unique_num in range(0, len(unique_courses)):
            unique_list = [row_num for row_num, info in lab_courses.items() if info == unique_courses[
                unique_num]]  # Generate a list of row numbers in the tuple for each course if the info is the same
            lab_groupings[unique_courses[unique_num]] = unique_list

        condensed_values = {}

        # Start condensing all of the identical courses into
        # TODO: Use Numpy to remove a for loop?
        # For every unique course
        for courses in unique_courses:
            lab_rows = lab_groupings[courses]  # Get the rows corresponding to this unique course
            course_data = []

            # For every column of data there is (recall range is not inclusive)
            for col in range(5, max_cols + 1):
                course_value = 0

                # For every row of this specific course, add up all of the data
                for row in lab_rows:
                    course_value = course_value + sheet_obj.cell(row=row, column=col).value
                    sheet_obj.cell(row=row, column=col).value = None
                course_data.append(course_value)
            condensed_values[courses] = course_data

            # Clear all of the condensed courses
            for row in lab_rows:
                for col in range(1, max_cols + 1):
                    sheet_obj.cell(row=row, column=col).value = None

        # Repopulate the bottom of the sheet with the new condensed values
        for course in unique_courses:
            # Placeholder Enrollment Code
            sheet_obj.cell(row=max_row, column=1).value = "Condensed Course"
            # TODO: put course name in above string

            condensed_numbers = condensed_values[course]

            # Repopulate the course name, professor, and time
            info_list = list(course)
            for col in range(0, 3):
                sheet_obj.cell(row=max_row, column=col + 2).value = info_list[col]

            # Repopulate the new values
            for col in range(5, max_cols + 1):
                sheet_obj.cell(row=max_row, column=col).value = condensed_numbers[col - 5]
            # Move onto the next row
            max_row += 1

        # Delete the blank rows
        # TODO: track the cells that you condensed into the first one in an array, and iterate though to potentially optimize code? Or can we delete when we condense courses instead of after?
        max_rows = sheet_obj.max_row
        # For every row
        row = 1
        while row != max_rows:
            # If the current cell is blank
            if sheet_obj.cell(row=row, column=1).value is None:
                # Create a blank array to track all sequential blank rows
                deletion_frame = []
                deletion_row = row
                # While the current row is blank
                # I made this while loop run instead of using an operator because
                while True:
                    if sheet_obj.cell(row=deletion_row, column=1).value is not None:
                        break
                    deletion_frame.append(deletion_row)
                    deletion_row = deletion_row + 1
                # Delete all blank rows using openpyxl's method
                sheet_obj.delete_rows(row, len(deletion_frame))

            max_rows = sheet_obj.max_row      # Is there a way to modify the for loop?
            row = row + 1

#        sheet_obj.protection.sheet = True
        # print("Done condensing lab courses in sheet " + str(sheet))

    # Create a sheet for the readme as the first sheet and add the readme text
    wb_obj.create_sheet("ReadMe")
    wb_obj.active = 3
    sheet_obj = wb_obj.active
    readme_cell = sheet_obj.cell(row=7, column=1)
    readme_cell.value = open("ReadMe.txt", "r").read()
    # alignment = Alignment(vertical="center", wrap_text=True)
    readme_cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet_obj.merge_cells("A7:K28")
    # sheet_obj.protection.sheet = True

    # Recalculate the percentages for condensed courses
    # TODO: Optimize code by saving what row number the condensed course was placed in a list, then iterate through that list instead of iterating through all rows
    print("Fixing percentages.")
    wb_obj.active = 1
    sheet_obj = wb_obj.active
    max_cols = sheet_obj.max_column
    max_rows = sheet_obj.max_row

    for row in range(1, max_rows+1):    # +1 is added here since range is not inclusive of the top and max_rows directly corresponds to the last row that has a value
        if sheet_obj.cell(row=row, column=1).value == "Condensed Course":
            # For every column starting from "E" to the end:
            for col in range(5, max_cols+1):
                # Access the first sheet's (course spaces left) value
                wb_obj.active = 0
                sheet_obj = wb_obj.active
                space_left = sheet_obj.cell(row=row, column=col).value

                # Access the third sheet's (ALl courses total space) value
                wb_obj.active = 2
                sheet_obj = wb_obj.active
                total_spaces = sheet_obj.cell(row=row, column=col).value

                # Compute percentage and change the value of the current cell in "All Courses % Space Left"
                wb_obj.active = 1
                sheet_obj = wb_obj.active

                print(str(row) + " " + str(col))
                print(space_left)
                print(total_spaces)
                print(max_cols)

                #if sheet_obj.cell(row=row, column=col).value is not None:
                sheet_obj.cell(row=row, column=col).value = space_left/total_spaces
                # If the percent actually means the percentage filled:
                # sheet_obj.cell(row=row, column=col).value = 1-(space_left/total_spaces)
                sheet_obj.cell(row=row, column=col).number_format = numbers.FORMAT_PERCENTAGE

    # Save the worksheet
    wb_obj.save(filename=condense_fname)
    wb_obj.save(filename=condense_fname)


def inject_macro(final_fname, vba_txt, button_txt):
    # This function takes in input of an absolute path, such as:
    # path = r"C:\Users\Dylan Vu\PycharmProjects\CARATS\Condensing Result.xlsm". Please note this code ONLY WORKS for Windows users!
    # and the name of a txt document containing the main VBA code in the same folder as this functions file:
    # VBA Code.txt
    # injects the main VBA code. This macro does not run.
    # Then, it injects vba code for a button:
    # VBA Button.txt
    # Runs the button vba macro, and then it then saves the Excel sheet specified in path.
    # Main resources used: https://pbpython.com/windows-com.html (Automating using pywin32),
    # https://stackoverflow.com/questions/19505676/use-python-to-write-vba-script (Injecting VBA code using Python)

    # Determine the path of the Excel file
    path = str(Path.cwd()) + r"/" + final_fname        # TODO: Potentially generalize this better?
    print("Injecting VBA code into sheet.")
    # Open the text file containing the VBA code
    vba_code = open(vba_txt, "r").read()

    # Use winpy32 to open up the specified Excel file in the background
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False           # Set to true if you want to see the file open
    ss = excel.Workbooks.Open(path)

    # Add the VBA code
    excel_module = ss.VBProject.VBComponents.Add(1)
    excel_module.CodeModule.AddFromString(vba_code)
    print("Creating Button.")
    # Inject button generation code in a new module
    button_code = open(button_txt, "r").read()
    excel_module = ss.VBProject.VBComponents.Add(1)
    excel_module.CodeModule.AddFromString(button_code)

    # Run the button VBA code. Single quotes around the file name are necessary since there are spaces.
    excel.Application.Run(final_fname + "!Module2.ButtonMacro")

    # Overwrite the original Excel file and save it
    excel.DisplayAlerts = False     # Disable the "Are you sure you want to overwrite this" prompt
    ss.SaveAs(path)
    excel.DisplayAlerts = True      # Enable the prompt again
    excel.Application.Quit()
